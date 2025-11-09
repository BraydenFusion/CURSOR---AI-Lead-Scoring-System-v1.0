"""Services for bulk lead import/export workflows."""

from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from uuid import UUID, uuid4

import pandas as pd
from fastapi import UploadFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.comments import Comment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session, joinedload

from ..database import SessionLocal
from ..models.activity import LeadActivity
from ..models.lead import Lead, LeadStatus
from ..models.note import LeadNote
from ..models.ai_scoring import LeadScore
from ..services.scoring_service import calculate_lead_score

try:
    # Prefer AI scoring when available
    from ..services.ai_scoring import calculate_overall_score
    except Exception:  # pragma: no cover - AI scoring optional dependency
    calculate_overall_score = None  # type: ignore


REQUIRED_COLUMNS: List[str] = ["name", "email", "phone", "company", "location", "source"]
MAX_IMPORT_ROWS = 10_000


class ImportErrorDetail(Dict[str, Any]):
    """Typed dict for import error details."""


def _load_dataframe_from_bytes(file_bytes: bytes, *, kind: str) -> pd.DataFrame:
    """Load dataframe from raw bytes supporting csv and excel."""

    if not file_bytes:
        raise ValueError("Uploaded file is empty.")

    buffer = io.BytesIO(file_bytes)
    if kind == "csv":
        df = pd.read_csv(buffer)
    elif kind == "excel":
        df = pd.read_excel(buffer)
    else:  # pragma: no cover - defensive
        raise ValueError("Unsupported file format.")

    if df.empty:
        raise ValueError("Uploaded file contains no data.")

    missing_columns = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

    if len(df) > MAX_IMPORT_ROWS:
        raise ValueError(f"Import limited to {MAX_IMPORT_ROWS} rows per upload.")

    return df


def _normalize_value(value: Any) -> Optional[str]:
    if pd.isna(value):
        return None
    if isinstance(value, str):
        value = value.strip()
    return str(value) if value not in (None, "", "nan") else None


def _prepare_metadata(row: pd.Series) -> Dict[str, Any]:
    metadata: Dict[str, Any] = {}
    company = _normalize_value(row.get("company"))
    if company:
        metadata["company"] = company
    for column in row.index:
        if column in REQUIRED_COLUMNS:
            continue
        metadata[column] = _normalize_value(row[column])
    return metadata


def _fetch_existing_emails(session: Session, emails: Sequence[str]) -> set[str]:
    if not emails:
        return set()
    results = session.query(Lead.email).filter(Lead.email.in_(emails)).all()
    return {email for (email,) in results}


def _create_lead_from_row(
    session: Session,
    row: pd.Series,
    user_id: UUID | None,
) -> Lead:
    metadata = _prepare_metadata(row)

    lead = Lead(
        id=uuid4(),
        name=_normalize_value(row["name"]) or "",
        email=_normalize_value(row["email"]) or "",
        phone=_normalize_value(row.get("phone")),
        source=_normalize_value(row.get("source")) or "import",
        location=_normalize_value(row.get("location")),
        created_by=user_id,
        _metadata=metadata,
        current_score=0,
        classification=None,
    )
    session.add(lead)
    session.flush()

    # Attempt AI scoring first, fallback to rules-based scoring.
    if calculate_overall_score:
        try:
            calculate_overall_score(lead.id, session)
        except Exception:  # pragma: no cover - fallback
            calculate_lead_score(lead.id, session)
    else:
        calculate_lead_score(lead.id, session)

    session.refresh(lead)
    return lead


def _validate_row(row_index: int, row: pd.Series, file_duplicates: set[str]) -> Tuple[bool, Optional[str]]:
    """Validate a row before inserting, returning tuple(is_valid, error_message)."""

    name = _normalize_value(row.get("name"))
    email = _normalize_value(row.get("email"))
    source = _normalize_value(row.get("source"))

    if not name:
        return False, "Name is required"
    if not email:
        return False, "Email is required"
    if "@" not in email:
        return False, "Email format invalid"
    if not source:
        return False, "Source is required"

    # Duplicate detection inside file
    if email in file_duplicates:
        return False, "Duplicate email within file"
    file_duplicates.add(email)
    return True, None


async def import_leads_from_csv(file: UploadFile, user_id: UUID | None) -> Dict[str, Any]:
    file_bytes = await file.read()
    df = _load_dataframe_from_bytes(file_bytes, kind="csv")
    return await _import_from_dataframe(df, user_id)


async def import_leads_from_excel(file: UploadFile, user_id: UUID | None) -> Dict[str, Any]:
    file_bytes = await file.read()
    df = _load_dataframe_from_bytes(file_bytes, kind="excel")
    return await _import_from_dataframe(df, user_id)


async def _import_from_dataframe(df: pd.DataFrame, user_id: UUID | None) -> Dict[str, Any]:
    session: Session = SessionLocal()
    summary: Dict[str, Any] = {
        "total_rows": len(df),
        "success_count": 0,
        "error_count": 0,
        "errors": [],
    }
    file_duplicates: set[str] = set()
    existing_emails = _fetch_existing_emails(session, [_normalize_value(e) or "" for e in df["email"].tolist()])

    try:
        for index, row in df.iterrows():
            excel_row_number = index + 2  # account for header
            is_valid, error_message = _validate_row(excel_row_number, row, file_duplicates)
            if not is_valid:
                summary["error_count"] += 1
                summary["errors"].append({"row": excel_row_number, "error": error_message})
                continue

            email = _normalize_value(row.get("email")) or ""
            if email in existing_emails:
                summary["error_count"] += 1
                summary["errors"].append({"row": excel_row_number, "error": "Lead with this email already exists"})
                continue

            try:
                lead = _create_lead_from_row(session, row, user_id)
                existing_emails.add(email)
                summary["success_count"] += 1
            except Exception as exc:  # pragma: no cover - defensive
                session.rollback()
                summary["error_count"] += 1
                summary["errors"].append({"row": excel_row_number, "error": str(exc)})
            finally:
                if summary["success_count"] % 100 == 0:
                    session.commit()

        session.commit()
        return summary
    finally:
        session.close()


def _apply_filters(query, filters: Dict[str, Any]) -> Any:
    classification = filters.get("classification")
    status = filters.get("status")
    source = filters.get("source")
    created_from = filters.get("created_from")
    created_to = filters.get("created_to")

    if classification and classification != "all":
        query = query.filter(Lead.classification == classification)
    if status and status != "all":
        try:
            query = query.filter(Lead.status == LeadStatus(status))
        except ValueError:
            pass
    if source:
        query = query.filter(Lead.source == source)
    if created_from:
        try:
            start = datetime.fromisoformat(created_from)
            query = query.filter(Lead.created_at >= start)
        except ValueError:
            pass
    if created_to:
        try:
            end = datetime.fromisoformat(created_to)
            query = query.filter(Lead.created_at <= end)
        except ValueError:
            pass
    return query


def _get_leads_with_filters(session: Session, filters: Dict[str, Any]) -> List[Lead]:
    query = session.query(Lead).options(
        joinedload(Lead.ai_scores),
        joinedload(Lead.activities),
        joinedload(Lead.notes),
    )
    query = _apply_filters(query, filters)
    return query.all()


def _serialize_leads(leads: Sequence[Lead]) -> List[Dict[str, Any]]:
    serialized: List[Dict[str, Any]] = []
    for lead in leads:
        latest_ai: Optional[LeadScore] = lead.ai_scores[-1] if lead.ai_scores else None

    serialized: List[Dict[str, Any]] = []
    for lead in leads:
        latest_ai: Optional[LeadScore] = lead.ai_scores[-1] if lead.ai_scores else None
        breakdown = {
            "engagement_score": getattr(latest_ai, "engagement_score", None),
            "buying_signal_score": getattr(latest_ai, "buying_signal_score", None),
            "demographic_score": getattr(latest_ai, "demographic_score", None),
        }
        serialized.append(
            {
                "id": str(lead.id),
                "name": lead.name,
                "email": lead.email,
                "phone": lead.phone,
                "source": lead.source,
                "location": lead.location,
                "current_score": lead.current_score,
                "classification": lead.classification,
                "status": lead.status.value if isinstance(lead.status, LeadStatus) else lead.status,
                "created_at": lead.created_at.isoformat(),
                "updated_at": lead.updated_at.isoformat(),
                "company": lead._metadata.get("company") if hasattr(lead, "_metadata") else None,
                "metadata": lead._metadata if hasattr(lead, "_metadata") else {},
                "score_breakdown": breakdown,
            }
        )
    return serialized


async def export_leads_to_csv(filters: Dict[str, Any]) -> bytes:
    session: Session = SessionLocal()
    try:
        leads = _get_leads_with_filters(session, filters)
        data = _serialize_leads(leads)
        if not data:
            return b""

        fields = list(data[0].keys())
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=fields)
        writer.writeheader()
        for row in data:
            writer.writerow(row)
        return buffer.getvalue().encode("utf-8")
    finally:
        session.close()


def _activities_dataframe(activities: Iterable[LeadActivity]) -> pd.DataFrame:
    records = [
        {
            "lead_id": str(activity.lead_id),
            "type": activity.activity_type,
            "details": activity.details,
            "created_at": activity.created_at,
        }
        for activity in activities
    ]
    return pd.DataFrame(records) if records else pd.DataFrame(columns=["lead_id", "type", "details", "created_at"])


def _notes_dataframe(notes: Iterable[LeadNote]) -> pd.DataFrame:
    records = [
        {
            "lead_id": str(note.lead_id),
            "author_id": str(note.created_by) if note.created_by else None,
            "content": note.content,
            "created_at": note.created_at,
        }
        for note in notes
    ]
    return pd.DataFrame(records) if records else pd.DataFrame(columns=["lead_id", "author_id", "content", "created_at"])


def _apply_classification_styles(worksheet: Worksheet, classification_column: int) -> None:
    hot_fill = PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid")
    warm_fill = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
    cold_fill = PatternFill(start_color="BFDBFE", end_color="BFDBFE", fill_type="solid")

    for row in worksheet.iter_rows(min_row=2):
        cell = row[classification_column]
        value = (cell.value or "").lower()
        if value == "hot":
            cell.fill = hot_fill
        elif value == "warm":
            cell.fill = warm_fill
        elif value == "cold":
            cell.fill = cold_fill


async def export_leads_to_excel(filters: Dict[str, Any]) -> bytes:
    session: Session = SessionLocal()
    try:
        lead_models = _get_leads_with_filters(session, filters)
        serialized_leads = _serialize_leads(lead_models)
        lead_ids = [lead.id for lead in lead_models]

        if lead_ids:
            activities = (
                session.query(LeadActivity)
                .filter(LeadActivity.lead_id.in_(lead_ids))
                .all()
            )
            notes = (
                session.query(LeadNote)
                .filter(LeadNote.lead_id.in_(lead_ids))
                .all()
            )
        else:
            activities = []
            notes = []

        workbook = Workbook()
        workbook.remove(workbook.active)

        # Leads sheet
        leads_sheet = workbook.create_sheet("Leads")
        if serialized_leads:
            leads_df = pd.DataFrame(serialized_leads)
            for row in dataframe_to_rows(leads_df, index=False, header=True):
                leads_sheet.append(row)
            classification_idx = list(leads_df.columns).index("classification")
            _apply_classification_styles(leads_sheet, classification_column=classification_idx)
        else:
            leads_sheet.append(["No lead data for selected filters"])

        # Activities sheet
        activities_sheet = workbook.create_sheet("Activities")
        activities_df = _activities_dataframe(activities)
        if not activities_df.empty:
            for row in dataframe_to_rows(activities_df, index=False, header=True):
                activities_sheet.append(row)
        else:
            activities_sheet.append(["No activities recorded"])

        # Notes sheet
        notes_sheet = workbook.create_sheet("Notes")
        notes_df = _notes_dataframe(notes)
        if not notes_df.empty:
            for row in dataframe_to_rows(notes_df, index=False, header=True):
                notes_sheet.append(row)
        else:
            notes_sheet.append(["No notes recorded"])

        for sheet in workbook.worksheets:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()
    finally:
        session.close()


def generate_import_template() -> bytes:
    """Generate Excel template with sample rows and guidance."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Lead Import Template"

    headers = [
        ("name", "Full name of the lead (required)"),
        ("email", "Unique email address (required)"),
        ("phone", "Primary phone number"),
        ("company", "Company or organization"),
        ("location", "City, state or region"),
        ("source", "Lead source e.g. Webinar, Website"),
    ]

    for idx, (header, description) in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.comment = None  # ensure reset
        cell.comment = Comment(description, "LeadScore AI")
        sheet.cell(row=2, column=idx, value={"name": "Alex Johnson", "email": "alex.johnson@example.com", "phone": "+1-555-123-4567", "company": "Acme Corp", "location": "New York, NY", "source": "Webinar"}[header])
        sheet.cell(row=3, column=idx, value={"name": "Jamie Lin", "email": "jamie.lin@example.com", "phone": "+1-555-987-6543", "company": "Northwind Traders", "location": "San Francisco, CA", "source": "Conference"}[header])

    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 24
    sheet.column_dimensions["E"].width = 22
    sheet.column_dimensions["F"].width = 18

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
"""Lead import and export services."""

from __future__ import annotations

import asyncio
import base64
import io
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Optional
from uuid import UUID, uuid4

import pandas as pd
from fastapi import UploadFile
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pandas import DataFrame
from sqlalchemy.orm import Session, joinedload

from ..database import SessionLocal
from ..models.assignment import LeadAssignment
from ..models.lead import Lead
from ..models.note import LeadNote
from ..models.activity import LeadActivity
from ..models.user import User, UserRole
from ..services.scoring_service import calculate_lead_score

EXPECTED_COLUMNS = ["name", "email", "phone", "company", "location", "source"]
IMPORT_BATCH_SIZE = 100


@dataclass
class ImportSummary:
    total_rows: int
    success_count: int
    error_count: int
    errors: List[Dict[str, Any]]
    preview: List[Dict[str, Any]]
    error_report: Optional[str] = None


def _clean_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    text = str(value).strip()
    if text.lower() in {"", "nan", "none"}:
        return None
    return text


def _read_dataframe_from_csv(contents: bytes) -> DataFrame:
    text_buffer = io.StringIO(contents.decode("utf-8-sig"))
    df = pd.read_csv(text_buffer, dtype=str).fillna("")
    return df


def _read_dataframe_from_excel(contents: bytes) -> DataFrame:
    bytes_buffer = io.BytesIO(contents)
    df = pd.read_excel(bytes_buffer, dtype=str).fillna("")
    return df


def _validate_columns(df: DataFrame) -> None:
    df.columns = [str(col).strip().lower() for col in df.columns]
    missing = [col for col in EXPECTED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")


def _collect_existing_emails(session: Session, emails: Iterable[str]) -> set[str]:
    normalized = {email.lower() for email in emails if email}
    if not normalized:
        return set()
    existing = (
        session.query(Lead.email)
        .filter(Lead.email.in_(list(normalized)))
        .all()
    )
    return {email.lower() for (email,) in existing}


def _calculate_latest_score_components(lead: Lead) -> Dict[str, Any]:
    breakdown = {
        "engagement": None,
        "buying_signals": None,
        "demographic_fit": None,
    }
    if lead.ai_scores:
        latest = max(lead.ai_scores, key=lambda score: score.scored_at)
        breakdown = {
            "engagement": latest.engagement_score,
            "buying_signals": latest.buying_signal_score,
            "demographic_fit": latest.demographic_score,
        }
    return breakdown


def _serialize_lead(
    lead: Lead,
    include_fields: Iterable[str],
) -> Dict[str, Any]:
    include = set(include_fields)
    metadata = lead._metadata or {}
    data: Dict[str, Any] = {}

    if not include or "basic" in include:
        data.update(
            {
                "id": str(lead.id),
                "name": lead.name,
                "email": lead.email,
                "source": lead.source,
                "classification": lead.classification,
                "status": lead.status.value if hasattr(lead.status, "value") else str(lead.status),
                "created_at": lead.created_at.isoformat(),
                "updated_at": lead.updated_at.isoformat(),
            }
        )

    if "contact" in include:
        data.update(
            {
                "phone": lead.phone,
                "location": lead.location,
                "company": metadata.get("company") or metadata.get("company_name"),
            }
        )

    if "score" in include:
        breakdown = _calculate_latest_score_components(lead)
        data.update(
            {
                "current_score": lead.current_score,
                "score_engagement": breakdown["engagement"],
                "score_buying_signals": breakdown["buying_signals"],
                "score_demographic_fit": breakdown["demographic_fit"],
            }
        )

    if "activities" in include:
        activity_summary = [
            f"{activity.activity_type} ({activity.points_awarded})"
            for activity in sorted(lead.activities, key=lambda a: a.timestamp)
        ]
        data["activities"] = "; ".join(activity_summary)

    if "notes" in include:
        note_summary = [
            f"[{note.note_type}] {note.content}"
            for note in sorted(lead.notes, key=lambda n: n.created_at)
        ]
        data["notes"] = "; ".join(note_summary)

    if "assignments" in include:
        assignment_summary = [
            f"{assignment.user.full_name if assignment.user else assignment.user_id} ({assignment.status})"
            for assignment in lead.assignments
        ]
        data["assignments"] = "; ".join(assignment_summary)

    return data


def _apply_lead_filters(query, filters: Dict[str, Any], current_user: Optional[User]) -> Any:
    if current_user:
        role = current_user.get_role_enum()
        if role == UserRole.SALES_REP:
            query = query.filter(Lead.created_by == current_user.id)

    classification = filters.get("classification")
    if classification and classification.lower() != "all":
        query = query.filter(Lead.classification == classification.lower())

    status = filters.get("status")
    if status:
        query = query.filter(Lead.status == status)

    source = filters.get("source")
    if source:
        query = query.filter(Lead.source == source)

    created_from = filters.get("created_from")
    if created_from:
        query = query.filter(Lead.created_at >= created_from)

    created_to = filters.get("created_to")
    if created_to:
        query = query.filter(Lead.created_at <= created_to)

    return query


def _process_import_dataframe(df: DataFrame, user_id: UUID) -> ImportSummary:
    _validate_columns(df)
    total_rows = len(df.index)
    summary = ImportSummary(
        total_rows=total_rows,
        success_count=0,
        error_count=0,
        errors=[],
        preview=df.head(5).to_dict(orient="records"),
    )

    session: Session = SessionLocal()
    try:
        normalized_emails = [_clean_str(email) for email in df["email"].tolist()]
        existing_emails = _collect_existing_emails(session, filter(None, normalized_emails))
        seen_emails: set[str] = set()
        created_leads: List[Lead] = []

        for index, row in df.iterrows():
            row_number = int(index) + 2  # include header offset
            name = _clean_str(row.get("name"))
            email = _clean_str(row.get("email"))
            phone = _clean_str(row.get("phone"))
            company = _clean_str(row.get("company"))
            location = _clean_str(row.get("location"))
            source = _clean_str(row.get("source")) or "Imported"

            if not name:
                summary.errors.append({"row_number": row_number, "email": email, "message": "Missing name"})
                continue
            if not email or "@" not in email:
                summary.errors.append({"row_number": row_number, "email": email, "message": "Invalid email"})
                continue

            email_key = email.lower()
            if email_key in seen_emails or email_key in existing_emails:
                summary.errors.append(
                    {"row_number": row_number, "email": email, "message": "Duplicate email"}
                )
                continue

            metadata: Dict[str, Any] = {}
            if company:
                metadata["company"] = company

            lead = Lead(
                id=uuid4(),
                name=name,
                email=email,
                phone=phone,
                source=source,
                location=location,
                created_by=user_id,
                _metadata=metadata,
            )

            session.add(lead)
            created_leads.append(lead)
            seen_emails.add(email_key)
            summary.success_count += 1

            if summary.success_count % IMPORT_BATCH_SIZE == 0:
                session.flush()
                for created_lead in created_leads:
                    try:
                        calculate_lead_score(created_lead.id, session)
                    except Exception:
                        # If scoring fails, continue importing remaining leads
                        pass
                session.commit()
                created_leads.clear()

        # Final flush and scoring for remaining leads
        session.flush()
        for created_lead in created_leads:
            try:
                calculate_lead_score(created_lead.id, session)
            except Exception:
                pass
        session.commit()

        summary.error_count = len(summary.errors)

        if summary.errors:
            error_df = pd.DataFrame(summary.errors)
            error_csv_bytes = error_df.to_csv(index=False).encode("utf-8-sig")
            summary.error_report = base64.b64encode(error_csv_bytes).decode("ascii")

        return summary
    finally:
        session.close()


async def import_leads_from_csv(file: UploadFile, user_id: UUID) -> Dict[str, Any]:
    """Import leads from a CSV file."""
    contents = await file.read()
    dataframe = await asyncio.to_thread(_read_dataframe_from_csv, contents)
    summary = await asyncio.to_thread(_process_import_dataframe, dataframe, user_id)
    return asdict(summary)


async def import_leads_from_excel(file: UploadFile, user_id: UUID) -> Dict[str, Any]:
    """Import leads from an Excel spreadsheet."""
    contents = await file.read()
    dataframe = await asyncio.to_thread(_read_dataframe_from_excel, contents)
    summary = await asyncio.to_thread(_process_import_dataframe, dataframe, user_id)
    return asdict(summary)


def _fetch_leads_for_export(filters: Dict[str, Any], include_fields: List[str], current_user: Optional[User]) -> List[Lead]:
    session: Session = SessionLocal()
    try:
        query = (
            session.query(Lead)
            .options(
                joinedload(Lead.activities),
                joinedload(Lead.notes),
                joinedload(Lead.assignments).joinedload(LeadAssignment.user),
                joinedload(Lead.ai_scores),
            )
        )
        query = _apply_lead_filters(query, filters, current_user)
        leads = query.all()
        return leads
    finally:
        session.close()


async def export_leads_to_csv(
    filters: Dict[str, Any],
    include_fields: Optional[List[str]] = None,
    current_user: Optional[User] = None,
) -> bytes:
    """Export filtered leads to CSV."""
    include_fields = include_fields or ["basic", "contact", "score"]

    def _to_csv() -> bytes:
        leads = _fetch_leads_for_export(filters, include_fields, current_user)
        rows = [_serialize_lead(lead, include_fields) for lead in leads]
        df = pd.DataFrame(rows)
        if df.empty:
            df = pd.DataFrame(columns=[field for field in include_fields])
        return df.to_csv(index=False).encode("utf-8-sig")

    return await asyncio.to_thread(_to_csv)


def _build_excel_workbook(
    leads: List[Lead],
    include_fields: List[str],
) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        lead_rows = [_serialize_lead(lead, include_fields) for lead in leads]
        leads_df = pd.DataFrame(lead_rows)
        if leads_df.empty:
            leads_df = pd.DataFrame(columns=["name", "email", "source"])
        leads_df.to_excel(writer, sheet_name="Leads", index=False)

        if "activities" in include_fields:
            activities = []
            for lead in leads:
                for activity in lead.activities:
                    activities.append(
                        {
                            "lead_email": lead.email,
                            "activity_type": activity.activity_type,
                            "points_awarded": activity.points_awarded,
                            "timestamp": activity.timestamp.isoformat(),
                        }
                    )
            activities_df = pd.DataFrame(activities)
            activities_df.to_excel(writer, sheet_name="Activities", index=False)

        if "notes" in include_fields:
            notes = []
            for lead in leads:
                for note in lead.notes:
                    notes.append(
                        {
                            "lead_email": lead.email,
                            "note_type": note.note_type,
                            "content": note.content,
                            "created_at": note.created_at.isoformat(),
                        }
                    )
            notes_df = pd.DataFrame(notes)
            notes_df.to_excel(writer, sheet_name="Notes", index=False)

        if "assignments" in include_fields:
            assignments = []
            for lead in leads:
                for assignment in lead.assignments:
                    assignments.append(
                        {
                            "lead_email": lead.email,
                            "assigned_to": assignment.user.full_name if assignment.user else str(assignment.user_id),
                            "status": assignment.status,
                            "assigned_at": assignment.assigned_at.isoformat(),
                            "is_primary": assignment.is_primary,
                        }
                    )
            assignments_df = pd.DataFrame(assignments)
            assignments_df.to_excel(writer, sheet_name="Assignments", index=False)

        workbook = writer.book
        leads_sheet = workbook["Leads"]

        header_font = Font(bold=True)
        for cell in leads_sheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        classification_column = None
        score_column = None
        for idx, cell in enumerate(leads_sheet[1], start=1):
            header = str(cell.value).lower()
            if header == "classification":
                classification_column = idx
            if header == "current_score":
                score_column = idx

        if classification_column:
            fills = {
                "hot": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                "warm": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                "cold": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
            }
            for row in range(2, leads_sheet.max_row + 1):
                value = leads_sheet.cell(row=row, column=classification_column).value
                if value:
                    fill = fills.get(str(value).lower())
                    if fill:
                        leads_sheet.cell(row=row, column=classification_column).fill = fill

        if score_column:
            percent_column = leads_sheet.max_column + 1
            leads_sheet.cell(row=1, column=percent_column).value = "Score Percent"
            for row in range(2, leads_sheet.max_row + 1):
                score_cell = f"{get_column_letter(score_column)}{row}"
                percent_cell = leads_sheet.cell(row=row, column=percent_column)
                percent_cell.value = f"=IFERROR(MIN(MAX({score_cell}/100,0),1),\"\")"
                percent_cell.number_format = "0.00%"

        for column_cells in leads_sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            column_letter = get_column_letter(column_cells[0].column)
            leads_sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    buffer.seek(0)
    return buffer.getvalue()


async def export_leads_to_excel(
    filters: Dict[str, Any],
    include_fields: Optional[List[str]] = None,
    current_user: Optional[User] = None,
) -> bytes:
    """Export filtered leads to Excel with formatting."""
    include_fields = include_fields or ["basic", "contact", "score", "activities", "notes", "assignments"]

    def _to_excel() -> bytes:
        leads = _fetch_leads_for_export(filters, include_fields, current_user)
        return _build_excel_workbook(leads, include_fields)

    return await asyncio.to_thread(_to_excel)


async def generate_import_template() -> bytes:
    """Generate an Excel template for lead imports."""
    def _build_template() -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"

        headers = [
            ("name", "Full name of the lead (required)"),
            ("email", "Unique email address (required)"),
            ("phone", "Phone number including country code (optional)"),
            ("company", "Company or organization (optional)"),
            ("location", "City/State or region (optional)"),
            ("source", "Lead source, e.g. Website, Referral (required)"),
        ]

        ws.append([header for header, _ in headers])
        ws.append(["Jane Smith", "jane@example.com", "+1-555-0100", "Acme Corp", "New York, NY", "Website"])
        ws.append(["Carlos Rodriguez", "carlos@example.org", "+34-555-2211", "Globex", "Madrid, Spain", "Outbound"])

        for idx, (_, description) in enumerate(headers, start=1):
            header_cell = ws.cell(row=1, column=idx)
            header_cell.font = Font(bold=True)
            header_cell.comment = Comment(description, "Lead System")
            ws.column_dimensions[get_column_letter(idx)].width = 24

        instructions = wb.create_sheet("Instructions")
        instructions["A1"] = "Lead Import Instructions"
        instructions["A1"].font = Font(bold=True, size=14)
        instructions["A3"] = "1. Do not modify the header row."
        instructions["A4"] = "2. Required columns: name, email, source."
        instructions["A5"] = "3. Email addresses must be unique across all leads."
        instructions["A6"] = "4. Maximum upload size: 10,000 rows."
        instructions["A7"] = "5. Supported formats: CSV (UTF-8) or Excel (.xlsx)."

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    return await asyncio.to_thread(_build_template)

